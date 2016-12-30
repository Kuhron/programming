import math, random
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"

def f(x):
    return math.log(1+abs(math.log(1+abs(random.random()*x*math.sin(x)/(1+math.cos(x))))))

for row in range(1, 40):
    ws1.append([i for i in map(f, [i for i in range(600)])])
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
print(ws3['AA10'].value)
wb.save("C:/Users/Wesley/Desktop/Programming/openpyxltest.xlsx")